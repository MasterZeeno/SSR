import os
import sys
import platform
from datetime import datetime
from PIL import Image as PILImage
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException

def set_wb(path, read_only=True, data_only=True):
    if not path.lower().endswith('.xlsx'):
        path += '.xlsx'
    if not os.path.isfile(path):
        sys.exit(f"File not found — {path}")
    try:
        return load_workbook(path, read_only=read_only, data_only=data_only)
    except (InvalidFileException, OSError) as e:
        sys.exit(f"Unable to load workbook — {e}")

# Constants
REPORT_TITLE = "SAFETY STATISTICS REPORT"
PROJECT_INFO = {
    "Date Range": None,
    "Name": "Construction of the New Senate Building (Phase II)",
    "Site": "Navy village, Fort Bonifacio, Taguig City",
    "Code": "PE-01-NSBP2-23"
}
START_ROW = 6
ADDNL_ROW = 53
BASE_WIDTH_PTS = 69

# Dimensions for columns A to N
# Columns F–K share the same width: 12.29
SHEET_DIMENS = dict(zip(
    [chr(i) for i in range(ord('A'), ord('N') + 1)],
    [0.67, 8, 7.71, 15.14, 10.43, *[12.29] * 6, 5.14, 2, 0.67]
))

FG = "002445"
FG_LIGHT = "00386C"
BG = "CEE7FF"
BG_LIGHT = "E2F1FF"
BG_LIGHTER = "F5FAFF"

SOURCE_FILE = "NSB-P2 SSR"
IMG_PATH = "hcclogo.png"
NEW_IMG_PATH = "resized.png"

wb = set_wb(SOURCE_FILE)
visible_sheets = [s for s in wb.worksheets if s.sheet_state == 'visible']
if not visible_sheets:
    sys.exit("No visible sheets found.")
ws = visible_sheets[-1]

# Workbook prep
temp_wb = Workbook()
temp_ws = temp_wb.active

# Helper functions
def points_to_pixels(points):
    return int(points * 96 / 72)

def to_datetime(value=None):
    date_format = "%b-%-d" if platform.system() != "Windows" else "%b-%#d"
    if isinstance(value, datetime):
        return value.strftime(date_format)
    try:
        parsed = datetime.fromisoformat(str(value))
        return parsed.strftime(date_format)
    except (ValueError, TypeError):
        return str(value)

def to_str(value=None, fallback=''):
    return str(value) if value is not None else fallback

def to_int(value=None, fallback=0):
    if not to_str(value).strip():
        return fallback
        
    try:
        return int(value)
    except (ValueError, TypeError):
        return fallback

def borderArray(border=None):
    thin = Side(border_style="thin", color=FG_LIGHT)
    border_map = {
        'all': Border(top=thin, right=thin, bottom=thin, left=thin),
        'top': Border(top=thin),
        'right': Border(right=thin),
        'bottom': Border(bottom=thin),
        'left': Border(left=thin),
        'top_bottom': Border(top=thin, bottom=thin),
        'left_right': Border(left=thin, right=thin)
    }
    return border_map.get(border, Border())

def resolve_column_letter(column=None, fallback=None):
    if isinstance(column, int) and 1 <= column <= 255:
        return get_column_letter(column)
    try:
        column = str(column).strip().upper()
        if column.isdigit():
            return get_column_letter(int(column))
        return column
    except Exception:
        return fallback or column

def resolve_column_integer(column=None, fallback=None):
    if isinstance(column, int) and 1 <= column <= 254:
        return column
    try:
        return column_index_from_string(str(column).strip().upper())
    except Exception:
        return fallback

def FORMAT_CELL(
    worksheet=temp_ws,
    start_row=0,
    start_column='B',
    end_row=0,
    end_column=None,
    value=None,
    font_size=12,
    bold=True,
    italic=False,
    horizontal_align="center",
    vertical_align="center",
    border='all',
    foreground=FG,
    background=None,
    number_format=None
):
    start_row = to_int(start_row)
    if start_row < 1:
        return None
    
    end_row = to_int(end_row)
    if end_row < start_row:
        end_row = start_row
        
    start_column = resolve_column_letter(start_column)
    end_column = resolve_column_letter(end_column)

    if end_column and end_column != 'NONE':
        worksheet.merge_cells(range_string=f"{start_column}{start_row}:{end_column}{end_row}")
    else:
        end_column = start_column

    cell = worksheet.cell(start_row, resolve_column_integer(start_column))

    cellValue = value
    if horizontal_align != "center":
        if horizontal_align == "right":
            cellValue = f"{value}  "
        else:
            cellValue = f"  {value}"
                
    cell.value = cellValue
    cell.font = Font(
        color=foreground, bold=bold, italic=italic,
        name='Arial', size=font_size
    )
    cell.number_format = number_format if number_format else "General"
    if background:
        fill_color = {
            'light': BG_LIGHT,
            'lighter': BG_LIGHTER
        }.get(background, BG)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=fill_color
        )
    
    cell.alignment = Alignment(
        horizontal=horizontal_align,
        vertical=vertical_align,
        indent=2
    )
    
    borderStyle = borderArray(border)
    if borderStyle:
        for row in worksheet.iter_rows(
            min_row=start_row, max_row=end_row,
            min_col=resolve_column_integer(start_column),
            max_col=resolve_column_integer(end_column)
        ):
            for cell in row:
                cell.border = borderStyle

def retrieve_value(start_row=0, start_column='B', end_row=0, end_column=None, dataType='str', ws=ws):
    start_row = to_int(start_row)
    if start_row < 1:
        return None
    start_row += ADDNL_ROW

    end_row = to_int(end_row)
    end_row = end_row + ADDNL_ROW if end_row >= 1 else start_row

    if end_row < start_row:
        end_row = start_row

    start_column = resolve_column_integer(start_column)
    end_column = resolve_column_integer(end_column) if end_column else start_column

    if end_column < start_column:
        end_column = start_column

    try:
        values = []
        for row in ws.iter_rows(
            min_row=start_row, max_row=end_row,
            min_col=start_column, max_col=end_column
        ):
            for cell in row:
                prefVal = {
                    'int': to_int(cell.value),
                    'datetime': to_datetime(cell.value)
                }.get(dataType, to_str(cell.value))
                values.append(prefVal)
            
        return values if len(values) > 1 else values[0] if values else None
    except Exception:
        return None

REPORT_DATE = retrieve_value(3, 'Q')
PREV_SUM_TBL = retrieve_value(6, 'R', 14, 'T', ws=visible_sheets[-2])

# Sheet setup
temp_ws.title = f"As of {REPORT_DATE}"
for col, width in SHEET_DIMENS.items():
    temp_ws.column_dimensions[col].width = width
temp_ws._images.clear()

# Handle logo resizing
IMG_HEIGHT = points_to_pixels(int(BASE_WIDTH_PTS * 0.69))
with PILImage.open(IMG_PATH) as img:
    IMG_WIDTH = int(IMG_HEIGHT / (img.height / img.width))
    img = img.resize((IMG_WIDTH, IMG_HEIGHT))
    img.save(NEW_IMG_PATH)

temp_ws.add_image(Image(NEW_IMG_PATH), "F1")

REF_CODE = retrieve_value(8, 'B')
PROJECT_INFO["Date Range"] = retrieve_value(10, 'D')
WEEK_DATA = {f"DA{k}S": retrieve_value(r,'E',r,'K',dataType='datetime') for k, r in {'Y': 18, 'TE': 19}.items()}
DEST_FILE = f"{SOURCE_FILE} as of {REPORT_DATE}.xlsx"

def sort_after_key(d, pivot_key, priority_key="EMD"):
    found = False
    before = {}
    after = {}

    for k, v in d.items():
        if found:
            after[k] = v
        else:
            before[k] = v
            if k == pivot_key:
                found = True

    # Extract the priority key if it exists in the "after" dict
    priority_item = {priority_key: after.pop(priority_key)} if priority_key in after else {}

    # Sort the rest
    after_sorted = dict(sorted(after.items()))

    # Merge the parts: before, priority_key (if present), and sorted rest
    return {**before, **priority_item, **after_sorted}

def SET_SHEET_DATA(ws=ws):
    KEYS = ["REGULAR", "OVERTIME"]
    blocks = {}
    current_block = None
    collecting = False
    block_data = {}

    for row in range(18, ws.max_row + 1):
        label = retrieve_value(row, 3).strip().upper()
        
        if any(label.startswith(p) for p in KEYS):
            # Start of a new block
            if current_block and block_data:
                blocks[current_block] = sort_after_key(block_data.copy(), "ADMIN")
            current_block = label.split()[0]  # Use REGULAR / OVERTIME as key
            block_data = {}
            collecting = False
        elif "TOTAL" in label and collecting:
            # End of current block
            blocks[current_block] = sort_after_key(block_data.copy(), "ADMIN")
            current_block = None
            block_data.clear()
            collecting = False
            if all(k in blocks for k in KEYS):
                break
        elif collecting and label:
            # Collect manpower row
            if "DATE" not in label and label not in block_data:
                block_data[label] = retrieve_value(
                    start_row=row,
                    start_column='E',
                    end_column='K',
                    dataType='int'
                )
        elif "DATE" in label:
            collecting = True
    
    return blocks

SHEET_DATA = SET_SHEET_DATA(ws)

# Title
start_column, end_column = 'B', 'M'
FORMAT_CELL(
    start_row=START_ROW, start_column=start_column,
    end_column=end_column, end_row=START_ROW + 1,
    value=REPORT_TITLE, font_size=20
)

START_ROW += 2
FORMAT_CELL(
    start_row=START_ROW, start_column=start_column,
    end_column=end_column, value=REF_CODE,
    font_size=9, italic=True, background=True,
    horizontal_align='right'
)

# Project info
START_ROW += 2
for key, val in PROJECT_INFO.items():
    is_date = "date" in key.lower()
    horizontal_align, font_size = 'left', 10
    
    FORMAT_CELL(
        start_row=START_ROW,
        start_column='B',
        end_column='C',
        value=key if is_date else f"Project {key}",
        bold=False,
        border=None,
        font_size=font_size,
        horizontal_align=horizontal_align
    )
    FORMAT_CELL(
        start_row=START_ROW,
        start_column='D',
        end_column='K',
        value=val,
        border='bottom',
        font_size=font_size,
        horizontal_align=horizontal_align
    )
    START_ROW += 1

# Manpower rows
SUB_TOTALS = {}
for C in ("POWER", "HOURS"):
    CLASS = f"I. MAN{C}"
    multiplier = 8 if 'H' in C else 1
    row_label = f"I{CLASS}" if 'H' in C else CLASS
    START_ROW += 2

    FORMAT_CELL(
        start_row=START_ROW,
        start_column='B', end_column='N',
        value=row_label, horizontal_align='left',
        border=None, background=True
    )

    START_ROW += 1
    for key, arr_val in SHEET_DATA.items():
        START_ROW += 1
        if 'H' in C:
            SUB_TOTALS[key] = f"=SUM(E{START_ROW + len(arr_val.keys()) + 2}:K{START_ROW + len(arr_val.keys()) + 2})"
        TOTAL_ARRFX = [
            f"=SUM({chr(i)}{START_ROW + 2}:{chr(i)}{START_ROW + len(arr_val.keys()) + 1})"
            for i in range(ord('E'), ord('K') + 1)
        ]
        TYPE = f"{key} {'(8am-5pm)' if key.startswith('R') else '(6pm-10pm)'}"
        for key, values in dict(zip(
            [TYPE, "DATE", *arr_val.keys(), "TOTAL"],
            [WEEK_DATA['DAYS'], WEEK_DATA['DATES'], *arr_val.values(), TOTAL_ARRFX]
        )).items():
            is_heading = key in [TYPE, "DATE", "TOTAL"]
            background = "light" if key in TYPE else "lighter" if is_heading else None
            font_size = 10
            FORMAT_CELL(
                start_row=START_ROW,
                start_column='C', end_column='D',
                value=key, bold=is_heading, italic=not is_heading,
                horizontal_align="center" if is_heading else "left",
                font_size=font_size, background=background
            )
            for i, col in enumerate([chr(i) for i in range(ord('E'), ord('K') + 1)]):
                FORMAT_CELL(
                    start_row=START_ROW, start_column=col,
                    bold=is_heading, horizontal_align="center",
                    font_size=font_size, background=background,
                    number_format=None if is_heading else "#,##0",
                    value=values[i] if is_heading else to_int(values[i]) * multiplier
                )
            START_ROW += 1
    
    if 'H' in C:
        for key, val in SUB_TOTALS.items():
            START_ROW += 1
            FORMAT_CELL(
                start_row=START_ROW, start_column='C',
                end_column='D', bold=True, horizontal_align="left",
                background=True, font_size=11, value=f"TOTAL {key}"
            )
            FORMAT_CELL(
                start_row=START_ROW, start_column='E',
                end_column='F', bold=True, horizontal_align="center",
                font_size=11, value=val
            )
    
# Save final workbook
temp_wb.save(DEST_FILE)

if os.path.exists(NEW_IMG_PATH):
    os.remove(NEW_IMG_PATH)
    
sys.exit(0)
