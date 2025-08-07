

def get_html_content(excel_file='../NSB-P2 SSR.xlsx', msgs=[]):
    import os, re
    from calendar import month_name, month_abbr
    from datetime import date, datetime
    from types import SimpleNamespace
    
    MONTH_MAP = {full: abbr for full, abbr in zip(month_name[1:], month_abbr[1:])}
    DATE_REGEX = r'\b(' + '|'.join(sorted(map(re.escape, MONTH_MAP), key=len, reverse=True)) + r')\b'
    
    for k, t in {'txt': str, 'num': (int, float), 'obj': dict, 'arr': (list, tuple, set), 'date': (date, datetime)}.items():
        globals()[f'is{k}'] = lambda v, t=t: isinstance(v, t)
        
    def fmt_num(n=1, f=1): return round(float(n), 2) if isnum(n) else f
    def fmt_str(s=None, f=''): return re.sub(r'\s+', ' ', str(s)).strip() or f
    def fmt_date(d=None, f=''): return re.sub(DATE_REGEX, lambda m: MONTH_MAP[m.group(0)], d or f)
    def fmt_cell(v=None, f=''): return f"{v:,.0f}" if isnum(v) else fmt_date(fmt_str(v,f))
    def fmt_key(k=None, f=''): return fmt_str(k).replace(' ', '_').lower() if k else f
    
    def xdelim(txt, d='()', rev=False):
        if not txt or len(d) != 2:
            return ''
        
        od, cd = map(re.escape, d)
        pattern = f'{od}(.*?){cd}'
        
        return " ".join(
            re.findall(pattern, txt) if not rev else
            re.sub(pattern, '', txt).split()
        ).strip()
        
    to_obj = lambda o: SimpleNamespace(**{fmt_key(k): to_obj(v) for k, v in o.items()}) if isobj(o) else [to_obj(i) for i in o] if isarr(o) else o
    obj_name = lambda o: next((k for k, v in inspect.currentframe().f_back.f_locals.items() if v is o), None)
    
    aliaser = lambda s, i='', x='': '' if not s else ''.join(c for c in str(s) if (c.isupper() or c.isdigit() or c in '-–—' or c in i) and c not in x)

    def get_data(excel_file=excel_file):
        from openpyxl import load_workbook as load
        from openpyxl.utils import column_index_from_string as cs
        
        wb = load(excel_file, read_only=True, data_only=True)
        ws = [s for s in wb.worksheets if s.sheet_state == "visible"][-1]
        
        rows, cols = range(59, 68), range(cs('R'), cs('T') + 1)
        details = {}

        for r in rows:
            if (o := ws.cell(r, 2).value) is not None and (k := fmt_key(o)):
                o, v = fmt_cell(o), fmt_cell(ws.cell(r, 4).value)
                if k.startswith("ref"):
                    details["ref_key"] = " ".join(o.split()[:-1])
                    details["ref_val"] = o.split()[-1]
                elif k.startswith("safe"):
                    details["title"] = o.title()
                    details["alias"] = aliaser(o.title())
                elif k.startswith("proj"):
                    details.setdefault("project", {})[
                        fmt_key(o.split()[1])
                    ] = v.title() if 'code' not in k and 'name' not in k else v
                    if "alias" not in details["project"]:
                        details["project"]["alias"] = aliaser(excel_file, ' ')
                elif k.startswith("date"):
                    details["full"] = "As of " + fmt_str(ws.cell(r, 4).value)
                    details["short"] = xdelim(v,rev=True)
                    details["weekdays"] = xdelim(v)
                else:
                    details[k] = v
                    
        company_data = {
            
        }
        
        zee_data = {
            "zee": {
                "name": "Jay Ar Adlaon Cimacio, RN", "position": "Occupational Health Nurse",
                "licenses": "License No.: 0847170", "website": "facebook.com/MasterZeeno",
                "assets": "SSR_HTML_EMAIL/assets"
            }
        }
        
        
        
        
        return to_obj({
            **details,
            "summary_table": [
                ['Description', *[ws.cell(58, c).value for c in cols]],
                *[
                    [ws.cell(r, 16).value, *[fmt_cell(ws.cell(r, c).value) for c in cols]]
                    for r in rows
                ]
            ],
            "company": {
                "name": "Hilmarc's Construction Corporation", "website": "hcc.com.ph",
                "address": "1835 E. Rodriguez Sr. Ave., Immaculate Conception, Quezon City",
                "licenses": "ISO 9001:2015 Certified | PCAB License No. 3886 AAA", "alias": "HCC",
                "copyleft": f"© 1977-{str(datetime.now().year)}. All rights reserved."
            },
            "zee": {
                "name": "Jay Ar Adlaon Cimacio, RN", "position": "Occupational Health Nurse",
                "licenses": "License No.: 0847170", "website": "facebook.com/MasterZeeno",
                "assets": f"{os.path.basename(os.getcwd())}/assets"
            }
        })
    
    REPORT = get_data()
    print(REPORT)
    # for timeline in REPORT.properties:
        # print('Timeline:', timeline)
    exit(0)
    
    COLORS = obj(
        fg="#002445", fg_lite="#0a66c2", fg_var="#60607b",
        bg="#f5faff", bg_dark="#f3f2f0"
    )
    SUBJECT, RADIUS, LANG = f"{PROJECT.alias} {REPORT.alias} as of {REPORT.date}", "border-radius:6px", 'lang="en" dir="ltr"'
    HTML_PROPS = obj(
        wrapper=f"""{LANG} style="background-color:{COLORS.bg_dark};margin:0px;padding:0px;border:none;outline:none;width:100%;height:auto;-webkit-text-size-adjust:100%;-ms-text-size-adjust:100%;font-family:Ubuntu, Helvetica, Arial, sans-serif;">""",
        valign='valign="middle" style="vertical-align:middle', border=f"border:0.169px solid {COLORS.fg}", div_content=f"""<div style="dir:ltr;text-align:left;word-wrap:break-word;color:{COLORS.fg};font-size:16px;line-height:1.5;font-weight:300""",
        container=f"{RADIUS};width:512px;max-width:512px;background-color:#fff;margin:8px auto;padding:0px 24px 24px"
    )
    
    def make_img(src='zee'):
        if src not in {'zee', 'company'}:
            return None
        defaults = f'{HTML_PROPS.valign};color:{COLORS.fg_var};height:auto;font-size:8px;line-height:1;text-decoration:none;outline:none;border:none'
        href, alt, size, filename = REPORT.zee.website, REPORT.zee.name, 64, 'cimacio'
        if src == 'company':
            href, alt, size, filename = REPORT.company.website, REPORT.company.name, 32, 'hcc-logo'
        else:
            defaults += ";border-radius:100%"
        return f"""<a href="https://{href}/" width="{size}" {defaults};display:inline-block;width:{size}px;max-width:{size}px;" target="_blank">
            <img src="https://raw.githubusercontent.com/MasterZeeno/SSR/refs/heads/main/{ZEE.assets_folder}/{filename}.png" alt="{alt}" {defaults};width:100%;max-width:100%;"></a>"""
    
    def hr(s=12, c=3, a='center', v=0.69, f=COLORS.fg_var):
        a = a if a in ('center', 'left', 'right') else 'center'
        c, s, v = fmt_num(c, 1), fmt_num(s, 12), fmt_num(v, 0.69)
        return f"""<p role="separator" {HTML_PROPS.valign};user-select:none;cursor:none;color:{f};font-size:{s}px;opacity:{v};text-align:{a};">{'&mdash;'*c}</p>"""
        
    def br(c=1):
        c = fmt_num(c, 1)
        return '<br aria-hidden="true">' * c
    
    def bold(text=None, revert=False):
        prop = f"font-weight:{300 if revert else 600}"
        return f"""<span style="{prop};">{text}</span>""" if text else prop
    
    data = []
    for r in range(58, 68):
        row = []
        for c in range(16, 21):
            if c != 17:
                row.append(fmt_cell(ws[f"{gl(c)}{r}"].value))
        data.append(row)
    wb.close()
    
    data[0][0] = f"As of {fmt_date(REPORT.date)}"
    
    summary_table = f"""<div style="padding:12px 0px;"><table width="100%" role="table" border="0" cellspacing="0" cellpadding="0" {HTML_PROPS.valign};border-collapse:collapse;width:100%;font-size:13px;color:{COLORS.fg};"><tbody>"""
    tag, consts = "td", f'{HTML_PROPS.valign};padding:3px 6px'
    for i, row in enumerate(data):
        row_html = "<tr>"
        for j, cell in enumerate(row):
            str_condition = True if i > 0 and j == 0 else False
            bgf_condition = True if i == 0 or (i in (8,9) and j == 3) else False
            align = "text-align:" + ("center" if (i == 0 and j == 0) else ("left" if j == 0 else "center"))
            bg = "background-color:" + (COLORS.bg if bgf_condition else "transparent")
            font_weight = bold() if bgf_condition else bold(revert=True)
            font_style = "font-style:" + ("italic" if str_condition else "normal")
            if str_condition:
                cell = "&nbsp;" * 3 + cell
            cell_html = f"""<{tag} align="{align}" {consts};{align};{HTML_PROPS.border};{bg};{font_weight};{font_style};">{cell}</{tag}>"""
            row_html += cell_html
        summary_table += row_html + "</tr>"
    
    return SUBJECT, f"""<!DOCTYPE html>
<html {LANG}>
<head>
    <meta charset="utf-8">
    <meta http-equiv="X-UA-Compatible" content="IE=edge">
    <title>{SUBJECT}</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
</head>
<body {HTML_PROPS.wrapper}
    <div {HTML_PROPS.wrapper}
        <div style="{HTML_PROPS.container};">
            <div style="padding:12px 24px 24px;">
                {HTML_PROPS.div_content};line-height:0;">
                    <p style="font-size:22px;{bold()};color:{COLORS.fg_lite}">{REPORT.name} ({REPORT.alias})</p>{build_header()}
                </div>{br(2)}{hr(6, 69)}
                {HTML_PROPS.div_content};padding-top:16px;">
                    <p style="{bold()};font-size:18px;">{msgs[0]}</p><p>{msgs[1]}</p><p>{msgs[2]}</p>
                    {summary_table}</tbody></table></div><p>{msgs[3]}{bold(msgs[4])}</p>
                </div>
                {HTML_PROPS.div_content};">
                    {hr(a='left', f=COLORS.fg)}<p>Best Regards,</p>
                    <div style="line-height:0;">
                        {make_img("zee")}
                        <div {HTML_PROPS.valign};display:inline-block;line-height:0.23;padding-left:8px;">
                            <p style="color:{COLORS.fg_lite};{bold()};font-size:16px;">{ZEE.name}</p>
                            <p style="font-size:14px;">{ZEE.position}</p>
                            <p style="font-size:12px;color:{COLORS.fg_var};">{ZEE.licenses}</p>
                        </div>
                    </div>
                </div>
            </div>
            {br(2)}
            {HTML_PROPS.div_content};{RADIUS};background-color:{COLORS.bg_dark};padding:12px 24px 24px;">
                <p style="color:{COLORS.fg_var};text-align:justify;font-size:12px;">
                    {bold("Disclaimer:")}
                    {br(2)}
                    This is an {bold("automated message.")} Please do not reply directly to this email.
                    {br(2)}
                    This email, including any attachments and previous correspondence in the thread, is {bold("confidential")} and intended solely for the designated recipient(s).
                    If you are not the intended recipient, you are hereby notified that any review, dissemination, distribution, printing, or copying of this message and its contents is strictly prohibited.
                    If you have received this email in error or have unauthorized access to it, please notify the sender immediately and {bold("permanently delete all copies")} from your system.
                    {br(2)}
                    The sender and the organization shall not be held liable for any unintended transmission of confidential or privileged information.
                    {br(3)}
                </p>
                <div style="color:{COLORS.fg_var};text-align:center;">
                    {hr(6, 69)}{br()}{make_img("company")}
                    <div {HTML_PROPS.valign};display:inline-block;padding:4px 0px;line-height:0;">
                        <p style="font-size:16px;{bold()};color:{COLORS.fg};">{COMPANY.name}</p>
                        <p style="font-size:9px;padding-bottom:6px;">{COMPANY.address}</p>
                    </div>
                    <p style="font-size:10px;">{COMPANY.licenses}{br()}{COMPANY.copyleft}</p>
                </div>
            </div>
        </div>
    </div>
</body>
</html>"""


get_html_content()