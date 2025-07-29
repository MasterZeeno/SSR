import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException

def set_wb(path, read_only=True, data_only=True):
    if not path.lower().endswith('.xlsx'):
        path += '.xlsx'
        
    if not os.path.isfile(path):
        print(f"Error: File not found — {path}")
        sys.exit(1)

    try:
        wb = load_workbook(filename=path, read_only=read_only, data_only=data_only)
        return wb
    except (InvalidFileException, OSError) as e:
        print(f"Error: Unable to load workbook — {e}")
        sys.exit(1)

INFO = {
    "Name": [8, "Construction of the New Senate Building (Phase II)"],
    "Site": [9, "Navy village, Fort Bonifacio, Taguig City"],
    "Code": [10, "PE-01-NSBP2-23"]
}

SCRIPT_TITLE = "SAFETY STATISTICS REPORT"
ADDNL_ROW = 53
FG = "002445"
FG_LIGHT = "00386C"
# 58AFFF
BG = "93CBFF"
BG_LIGHT = "E2F1FF"
# BG_LIGHTEST = "E2F1FF"

SOURCE_FILE = "NSB-P2 SSR"
TOP_TEMPLATE_FILE = f"{SOURCE_FILE} - TEMPLATE"

try:
    wb = set_wb(SOURCE_FILE)
    visible_sheets = [sheet for sheet in wb.worksheets if sheet.sheet_state == 'visible']

    if not visible_sheets:
        print("No visible sheets found.")
        sys.exit(1)

    ws = visible_sheets[-1]
    
    report_date = ws[f"Q{ADDNL_ROW + 3}"].value
    ref_code = ws[f"B{ADDNL_ROW + 8}"].value
    date_range = ws[f"D{ADDNL_ROW + 10}"].value
    
    DEST_FILE = f"{SOURCE_FILE} as of {report_date}.xlsx"
    
    def to_str(s):
        return str(s) if s is not None else ''
        
    def to_int(s):
        try:
            return int(s)
        except (ValueError, TypeError):
            return 0

    MANPOWER_LIST = {}

    start = False
    for r in range(69, ws.max_row + 1):
        val = to_str(ws.cell(r, 3).value).strip().upper()

        if start:
            if 'TOTAL' not in val:
                if val:
                    data = []
                    for c in range(5, 12):
                        data.append(to_int(ws.cell(r, c).value))
                        
                    MANPOWER_LIST[val] = data
            else:
                break

        if 'DATE' in val:  # Use .upper() for robustness
            start = True

    # print(MANPOWER_LIST)
    
    temp_wb = set_wb(TOP_TEMPLATE_FILE, False, False)
    temp_ws = temp_wb.active
    temp_ws.title = f"As of {report_date}"
    
    thin = Side(border_style="thin", color=FG_LIGHT)
    
    def borderArray(border):
        if border == 'all':
            return Border(top=thin, right=thin, bottom=thin, left=thin)
        elif border == 'top':
            return Border(top=thin)
        elif border == 'right':
            return Border(right=thin)
        elif border == 'bottom':
            return Border(bottom=thin)
        elif border == 'left':
            return Border(left=thin)
        elif border == 'top_bottom':
            return Border(top=thin, bottom=thin)
        elif border == 'left_right':
            return Border(left=thin, right=thin)
        else:
            return Border()
    
    def res_col(s=None):
        if not s:
            return None
        elif isinstance(s, int):
            return get_column_letter(s)
        else:
            return str(s).strip().upper()
    
    def fmt_cell(row=None, max_row=None, sht=temp_ws, sc=2, ec=None, val=None, fz=12, bold=True, italic=False, ha="center", va="center", border='all', bg=None):
        row = to_int(row)
        if row < 1:
            return
            
        max_row = to_int(max_row)
        if max_row < row:
            max_row = row
            
        sc = res_col(sc)
        ec = res_col(ec)
            
        if ec:
            sht.merge_cells(range_string=f"{sc}{row}:{ec}{max_row}")
            
        cell = sht.cell(row, column_index_from_string(sc))
        
        if val:
            if ha != "center":
                val = f"  {val}"
            cell.value = val
            cell.font = Font(color=FG, bold=bold, italic=italic, name='Arial', size=fz)
        
        if bg:
            if bg == 'light':
                fgColor = BG_LIGHT
            else:
                fgColor = BG
                
            cell.fill = PatternFill(fill_type="solid", fgColor=fgColor)
        
        cell.border = borderArray(border)
        cell.alignment = Alignment(horizontal=ha, vertical=va)
        
    fmt_cell(3, 4, sc='b', ec='m', val=SCRIPT_TITLE, fz=20)
    fmt_cell(5, sc='b', ec='m', val=f"{ref_code}  ", fz=9, italic=True, ha='right', bg=True)
    
    INFO["Date Range"] = [7, date_range]
    for key, val in INFO.items():
        if not "Date" in key:
            key = f"Project {key}"
        fmt_cell(val[0], sc='b', ec='c', bold=False, val=key, ha='left', border=None, fz=10)
        fmt_cell(val[0], sc='d', ec='k', val=val[1], ha='left', border='bottom', fz=10)
        
    KEYS_LEN = len(MANPOWER_LIST.keys())
    START_ROW = 13
    for CLASS in ("POWER", "HOURS"):
        ROW_KEY = f"I. MAN{CLASS}"
        if CLASS == "HOURS":
            MULTIPLIER = 8
            ROW_KEY = f"II. MAN{CLASS}"
            START_ROW += 2
        else:
            MULTIPLIER = 1
        
        fmt_cell(START_ROW, sc='b', ec='n', val=ROW_KEY, ha='left', border=None, bg = "light")
        START_ROW += 1
        for TYPE in ("REGULAR (8am-5pm)", "OVERTIME (6pm-10pm)"):
            START_ROW += 1
            for SUBTYPE in [TYPE, "DATE", *MANPOWER_LIST.keys(), "TOTAL"]:
                if SUBTYPE in [TYPE, "DATE", "TOTAL"]:
                    is_bold = True
                    ha = "center"
                    if SUBTYPE == TYPE:
                        bg = True
                    else:
                        bg = "light"
                else:
                    is_bold = False
                    ha = "left"
                    bg = None
                    
                fmt_cell(START_ROW, sc='c', ec='d', val=SUBTYPE, bold=is_bold, fz=10, ha=ha, bg=bg)
                START_ROW += 1
                
        # for r in range(13, temp_ws.max_row + 1):
            # if ROW_KEY in to_str(temp_ws.cell(r, 2).value).strip().upper():
                
                # START_ROW = r + 4
                # SUBTYPE_OBJ = {}
                # for SUBTYPE in ("REGULAR", "OVERTIME"):
                    # if SUBTYPE == "OVERTIME":
                        # START_ROW = START_ROW + KEYS_LEN + 5
                        # if CLASS == "HOURS":
                            # MULTIPLIER = 2
                    
                    # SUBTYPE_OBJ[SUBTYPE] = [
                        # MULTIPLIER,
                        # START_ROW
                    # ]
                    
                    # temp_ws.insert_rows(START_ROW, KEYS_LEN)
                    # temp_wb.save(DEST_FILE)
                    # sys.exit(0)
                    # temp_wb = set_wb(DEST_FILE, False, False)
                    # temp_ws = temp_wb.active
                    
                # ROW_DATA[ROW_KEY] = SUBTYPE_OBJ
                
                    # for key, values in MANPOWER_LIST.items():
                        # temp_ws.cell(START_ROW, 3).value = key
                        # for idx, val in enumerate(values):
                            # cell = temp_ws.cell(START_ROW, idx + 5)
                            # cell.value = str(val * MULTIPLIER)
                            # cell.border = border
                        # START_ROW += 1
                        
    temp_wb.save(DEST_FILE)
except Exception as e:
    print(f"Error: {e}")
    sys.exit(1)
   