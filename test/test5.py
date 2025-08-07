
import os, re
from calendar import month_name, month_abbr
from datetime import date, datetime
from types import SimpleNamespace

MONTH_MAP = {full: abbr for full, abbr in zip(month_name[1:], month_abbr[1:])}
DATE_REGEX = r'\b(' + '|'.join(sorted(map(re.escape, MONTH_MAP), key=len, reverse=True)) + r')\b'

for k, t in {'txt': str, 'num': (int, float), 'obj': dict, 'arr': (list, tuple, set), 'date': (date, datetime)}.items():
    globals()[f'is{k}'] = lambda v, t=t: isinstance(v, t)
    
def fmt_num(n=1, f=1): return round(float(n), 2) if isnum(n) else f
def fmt_str(s=None, f=''): return re.sub(r'\s+', ' ', str(s)).strip() or f
def fmt_date(d=None, f=''): return re.sub(DATE_REGEX, lambda m: MONTH_MAP[m.group(0)], d or f)
def fmt_cell(v=None, f=''): return f"{v:,.0f}" if isnum(v) else fmt_date(fmt_str(v,f))
def fmt_key(k=None, f=''): return fmt_str(k).strip().replace(' ', '_').lower() if k else f

to_obj = lambda o: SimpleNamespace(**{fmt_key(k): to_obj(v) for k, v in o.items()}) if isobj(o) else [to_obj(i) for i in o] if isarr(o) else o
obj_name = lambda o: next((k for k, v in inspect.currentframe().f_back.f_locals.items() if v is o), None)


class Details:
    def __init__(self):
        self.details = {}

    def set_detail(self, key, original, value):
        self.details[key] = {
            'original': original,
            'value': value
        }

    def process_rows(self, ws, rows):
        for r in rows:
            if (o := ws.cell(r, 2).value) and (k := self.fmt_key(o)):
                o_fmt = self.fmt_cell(o)
                v_fmt = self.fmt_cell(ws.cell(r, 4).value)
                self.set_detail(k, o_fmt, v_fmt)

    def fmt_key(self, val):
        # You can customize this as needed
        return str(val).strip().lower().replace(" ", "_")

    def fmt_cell(self, val):
        return str(val).strip() if val is not None else ""

    def print(self):
        for key, val in self.details.items():
            print(f"{key}: {val['value']}")
            # print(f"{key}: {val['original']} → {val['value']}")
            
from openpyxl import load_workbook as load
from openpyxl.utils import column_index_from_string as cs
        
excel_file = '../NSB-P2 SSR.xlsx'

wb = load(excel_file, read_only=True, data_only=True)
ws = [s for s in wb.worksheets if s.sheet_state == "visible"][-1]

rows, cols = range(59, 68), range(cs('R'), cs('T') + 1)
details = Details()
details.process_rows(ws, rows)
details.print()