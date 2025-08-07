from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gl
import sys
import os
import re

def toint(s):
    s = re.sub(r'\D', '', str(s))
    try:
        return int(s)
    except (ValueError, TypeError):
        return 0
        
def format_cell(cell):
    if isinstance(cell, (int, float)):
        return f"{cell:,.0f}"
    return str(cell).strip() if cell is not None else ''

if len(sys.argv) == 2:
    path = sys.argv[1]
else:
    path = 'PE-01-NSBP2-23 SSR'

if not path.lower().endswith('.xlsx'):
    path += '.xlsx'
if not os.path.isfile(path):
    sys.exit(f"File not found — {path}")

wb = load_workbook(path, read_only=True, data_only=True)
shts = [s for s in wb.worksheets if s.sheet_state == 'visible']
if not shts:
    sys.exit(1)
bws = shts[-2]
ws = shts[-1]

prev_manpwr, prev_manhrs, pres_manpwr, pres_manhrs = [
    toint(sht[coord].value)
    for sht in (bws, ws)
    for coord in ('S66', 'T67')
]

report_date, manpwr_header, manhrs_header, highest_manpwr = [
    ws[coord].value
    for coord in ('Q56', 'P66', 'P67', 'T66')
]

data = []
for r in range(58, 66):
    row = []
    for c in range(16, 21):
        if c != 17:
            row.append(ws[f"{gl(c)}{r}"].value)
    data.append(row)
wb.close()

data[0][0] = report_date
data.append([
    manpwr_header, prev_manpwr, pres_manpwr,
    max(prev_manpwr, pres_manpwr, highest_manpwr)
])
data.append([
    manhrs_header, prev_manhrs,
    pres_manhrs - prev_manhrs,
    pres_manhrs
])

# Generate HTML table rows with alignment classes
table_rows = ""
color = "0,97,186"
font_family = "-apple-system,system-ui,BlinkMacSystemFont,'Segoe UI',Roboto,'Helvetica Neue','Fira Sans',Ubuntu,Oxygen,'Oxygen Sans',Cantarell,'Droid Sans','Apple Color Emoji','Segoe UI Emoji','Segoe UI Emoji','Segoe UI Symbol','Lucida Grande',Helvetica,Arial,sans-serif"
table_attrs = """width="100%" role="presentation" valign="top" border="0" cellspacing="0" cellpadding="0"""
font_defaults = f"font-family:{font_family};color:rgb({color});text-align:left;word-spacing:normal;letter-spacing:normal"
default_styles = f"border:0.0123em solid rgb({color});padding:0.23em;"
for i, row in enumerate(data):
    tag = "th" if i == 0 else "td"
    row_html = "<tr>"
    for j, cell in enumerate(row):
        alignment = "center" if (i == 0 and j == 0) else ("left" if j == 0 else "center")
        
        if i == 0:
            addtnl = f"background-color:rgba({color},.069);"
        elif i > 0 and j == 0:
            cell = f"&nbsp;&nbsp;{cell}"
            addtnl = "font-style:italic;"
        else:
            addtnl = ""
            
        cell_html = f'<{tag} align="{alignment}" style="text-align:{alignment};{default_styles}{addtnl}">{format_cell(cell)}</{tag}>'
        row_html += cell_html
    row_html += "</tr>"
    table_rows += row_html

# Full HTML document
html_content = f"""<div dir="ltr" style="background:0 0;margin:0;padding:0;border:0 none transparent;outline:0 none transparent;width:100%;height:auto;box-sizing:border-box">
  <table {table_attrs}" style="font-size:16px;max-width:532px;min-width:300px;width:96.69%;box-sizing:border-box">
    <tbody>
      <tr>
        <td align="left">
          <table {table_attrs}" style="{font_defaults};font-size:clamp(.5em,2.353vw + .059em,1em);line-height:clamp(.4em,6.588vw + -.835em,1.8em);padding:0 clamp(1.125em,3.882vw + .397em,1.95em) 0;box-sizing:border-box">
            <tbody>
              <tr>
                <td>
                  <div dir="ltr" style="margin:clamp(1.125em,3.882vw + .397em,1.95em) auto 0;font-size:0.96em;box-sizing:border-box">
                    <h3>Good day, everyone!</h3>
                    <p>Please find the attached updated <b>Safety Statistics Report (SSR)</b>
                      for Project Code: <b>PE-01-NSBP2-23&nbsp;&ndash;&nbsp;Construction of the New Senate Building (Phase II).</b>
                    </p>
                    <p>
                      Alternatively, for your convenience, a brief summary is provided in the table below:
                    </p>
                    <table {table_attrs}" style="border-collapse:collapse;width:100%;{font_defaults};font-size:0.9em;box-sizing:border-box">
                        {table_rows}
                    </table>
                    <p>
                      Thank you, and as always&mdash;<b>Safety First!</b>&nbsp;👊
                    </p>
                  </div>
                </td>
              </tr>
              <tr>
                <td>
                  <div style="border:0 none transparent;border-top:.032em dashed rgba({color},.69);width:96%;margin:clamp(1.125em,3.882vw + .397em,1.95em) auto;box-sizing:border-box"></div>
                  <p>Best regards,</p>
                </td>
              </tr>
              <tr>
                <td align="center">
                  <a href="https://www.hcc.com.ph/" style="display:block;text-decoration:none" target="_blank">
                    <img role="presentation" src="https://raw.githubusercontent.com/MasterZeeno/Repository/main/zee-signature.png" alt="Jay Ar Adlaon Cimacio, RN, OHN" width="100%" height="auto" style="border:none;outline:none;text-decoration:none;display:block;max-width:100%;height:auto;border:none;box-sizing:border-box">
                  </a>
                </td>
              </tr>
            </tbody>
          </table>
        </td>
      </tr>
    </tbody>
  </table>
</div>"""

# Save to HTML file
output_path = "body.html"
with open(output_path, "w", encoding="utf-8") as f:
    f.write(html_content)
print(report_date)
