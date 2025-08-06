

def get_html_content(excel_file=None):
    import os, re
    from types import SimpleNamespace as obj
    from data_holder import PROJECT_DATA, MONTH_MAP, DATE_REGEX, xdelim
    
    for k, t in {'txt': str, 'num': (int, float), 'obj': dict, 'arr': (list, tuple, set, range)}.items():
        globals()[f'is{k}'] = lambda v, t=t: isinstance(v, t)
    
    to_obj = lambda o: obj(**{fmt_key(k): to_obj(v) for k, v in o.items()}) if isobj(o) else [to_obj(i) for i in o] if isarr(o) else o
    obj_name = lambda o: next((k for k, v in inspect.currentframe().f_back.f_locals.items() if v is o), None)
    
    aliaser = lambda s, i='', x='': '' if not s else ''.join(c for c in str(s) if (c.isupper() or c.isdigit() or c in '-–—' or c in i) and c not in x)
    fmt_num  = lambda n, f=1: round(float(n), 2) if isnum(n) else f
    fmt_str  = lambda s, f='', l=False: (v := re.sub(r'\s+', ' ', str(s)).strip()) and (v.lower() if l else v) if s else f
    fmt_key  = lambda k, f='', l=True, r=True: ((s := re.match(r"[A-Za-z ]*", fmt_str(k, f, l)).group().strip()) and (s.replace(' ', '_') if r else s)) if k else f
    fmt_cell = lambda v, f='': (f"{v:,.0f}" if isnum(v) else fmt_str(v)) if v else f
    fmt_date = lambda d, f='': re.sub(DATE_REGEX, lambda m: MONTH_MAP[m.group(0)], fmt_str(d,f).title()) if d else f
        
    
    
    def get_data(excel_file=excel_file):
        excel_file = excel_file or PROJECT_DATA.get("excel_file", None)
        if excel_file is None:
            exit(1)
            
        from openpyxl import load_workbook
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        ws = [s for s in wb.worksheets if s.sheet_state == "visible"][-1]
        REPORT_DETAILS, ROWS, COLS, PKEYS = {}, range(59, 68), range(18, 21), ('ref', 'date', 'project', 'safety')
        
        has = lambda s, r: fmt_key(s).startswith(PKEYS[r]) if isnum(r) and 0 <= r < len(PKEYS) else any(fmt_key(s).startswith(PKEYS[i]) for i in r if isnum(i) and 0 <= i < len(PKEYS)) if isarr(r) else False
        
        HEADER_OBJ = {}
        
        def add_header_entry(header_obj=HEADER_OBJ, pkey=None, pval=None, return_html=False):
            hkeys = ('headers', 'labels', 'contexts')
        
            if hkeys[0] not in header_obj:
                header_div = '<div valign="middle" style="vertical-align:middle;display:inline-block;font-size:15px;color:'
                clrs = [PROJECT_DATA["colors"].get(f"fg{k}", "#333") for k in ("_var", "")]
                header_obj[hkeys[0]] = {
                    hkeys[1]: [
                        f'{header_div}{clrs[0]};">'
                    ],
                    hkeys[2]: [
                        f'{header_div}{clrs[1]};font-weight:600;">'
                    ]
                }
        
            def append_entry(header_obj, pkey, pval, wrap=True):
                p_wrap = lambda x: f'<p style="line-height:0.5;">{x}</p>'
                for i, v in enumerate((pkey, pval), 1):
                    header_obj[hkeys[0]].setdefault(hkeys[i], []).append(p_wrap(v) if wrap else v)
            
            if all([v is not None for v in [header_obj, pkey, pval]]):
                append_entry(header_obj, pkey, pval)
                
                if len(header_obj[hkeys[0]][hkeys[1]]) == 3 or "date" in fmt_str(pkey,l=True):
                    append_entry(header_obj, *['<br aria-hidden="true">'] * 2, wrap=False)
            
            if return_html and len(header_obj[hkeys[0]][hkeys[1]]) > 0:
                return fmt_str(''.join(''.join(header_obj[hkeys[0]][key]) + '</div>' for key in (hkeys[1], hkeys[2]))).replace('\n', '')
                                
            
            

                
  
        for r in ROWS:
            if (o := ws.cell(r, 2).value) is not None and (k := fmt_key(o)):
                o, v = fmt_cell(o), fmt_str(ws.cell(r, 4).value)
                if has(k, range(0,3)):
                    if has(k, [0, 1]): # ref / date
                        if has(k,0):
                            idx = 0
                            pkey = fmt_key(" ".join(p:=o.split()[:-1]),l=False,r=False).title()
                            pval = o.split()[-1].upper()
                        else:
                            idx = 1
                            pkey = o.title()
                            pval = v.title()
                            
                        REPORT_DETAILS[PKEYS[idx]] = {
                            "key": pkey,
                            "value": pval
                        }
                        
                    if has(k, 2): # project
                        pkey = fmt_key(o.split()[1])
                        if pkey.startswith(("code", "name")):
                            pval = v
                        else:
                            pval = v.title()
                            
                        REPORT_DETAILS.setdefault(PKEYS[2], {})[pkey] = pval
                        pkey = f"{PKEYS[2].title()} {pkey.title()}"
                    
                    add_header_entry(HEADER_OBJ, pkey, pval)
          
                elif has(k, 3): # safe
                    REPORT_DETAILS["title"] = o.title()
                    REPORT_DETAILS["alias"] = aliaser(o.title())
                else:
                    REPORT_DETAILS[k] = fmt_cell(v)
        
        
        def html_cell(value, align="center", bg="#f5faff", weight="600", italic=False):
            return (
                f'<td align="{align}" valign="middle" '
                f'style="vertical-align:middle;padding:3px 6px;text-align:{align};'
                f'border:0.169px solid {PROJECT_DATA["colors"]["fg"]};background-color:{bg};'
                f'font-weight:{weight};font-style:{"italic" if italic else "normal"};">{value}</td>'
            )
        
        def html_row(cells, is_header=False):
            return "<tr>" + "".join([
                html_cell(
                    cell,
                    align="left" if i == 0 else "center",
                    bg="#f5faff" if is_header else "transparent",
                    weight="600" if is_header else "300",
                    italic=(i == 0 and not is_header)
                )
                for i, cell in enumerate(cells)
            ]) + "</tr>"
        
        def generate_summary_table(ws, COLS, ROWS):
            rows_html = [
                html_row(["Description", *[ws.cell(58, c).value for c in COLS]], is_header=True),
                *[
                    html_row([
                        ws.cell(r, 16).value,
                        *[fmt_cell(ws.cell(r, c).value, 0) for c in COLS]
                    ])
                    for r in ROWS
                ]
            ]
            
            return (
                '<table width="100%" role="table" border="0" cellspacing="0" cellpadding="0" valign="middle" '
                'style="vertical-align:middle;border-collapse:collapse;width:100%;font-size:13px;color:#002445;">\n'
                + "\n".join(rows_html) +
                '\n</table>'
            )
        
        
        # def html_cell(value, align="center", bg="#f5faff", weight="600", italic=False):
            # return (
                # f'<td align="{align}" valign="middle" '
                # f'style="vertical-align:middle;padding:3px 6px;text-align:{align};'
                # f'border:0.169px solid #002445;background-color:{bg};'
                # f'font-weight:{weight};font-style:{"italic" if italic else "normal"};">{value}</td>'
            # )
        
        # def html_row(cells, is_header=False):
            # return "<tr>" + "".join([
                # html_cell(
                    # cell,
                    # align="left" if i == 0 else "center",
                    # bg="#f5faff" if is_header else "transparent",
                    # weight="600" if is_header else "300",
                    # italic=(i == 0 and not is_header)
                # )
                # for i, cell in enumerate(cells)
            # ]) + "</tr>"
        
        # summary_rows = "".join([
            # html_row(["Description", *[ws.cell(58, c).value for c in COLS]], is_header=True),
            # *[
                # html_row([
                    # ws.cell(r, 16).value,
                    # *[fmt_cell(ws.cell(r, c).value, 0) for c in COLS]
                # ])
                # for r in ROWS
            # ]
        # ])
        
        
        return {
            **REPORT_DETAILS,
            **PROJECT_DATA,
            "headers": add_header_entry(return_html=True),
            "summary_table": generate_summary_table(ws, COLS, ROWS)
        }
            # "summary": [
                # ["Description", *[ws.cell(58, c).value for c in COLS]],
                # *[[ws.cell(r, 16).value, *[fmt_cell(ws.cell(r, c).value, 0) for c in COLS]] for r in ROWS]
            # ],
        
        # return to_obj({
            # **REPORT_DETAILS, **PROJECT_DATA,
            # "summary": [
                # ["Description", *[ws.cell(58, c).value for c in COLS]],
                # *[[ws.cell(r, 16).value, *[fmt_cell(ws.cell(r, c).value, 0) for c in COLS]] for r in ROWS]
            # ],
        # })
   
        
    # DUMMY = get_data()
    
    
    # print(REPORT)
    # exit(0)
    # for i, (k, v) in enumerate(REPORT.headers):
        # print('index:', i, '\nkey:', k, '\nvalue:', v, '\n')
    
    # import json
    # with open("output.json", "w", encoding="utf-8") as f:
        # json.dump(DUMMY, f, ensure_ascii=False, indent=4)
    # print('Json data saved.')
    
    # for key, val in [
        # [REPORT.ref_key, REPORT.ref_val],
        # [REPORT.date.key, REPORT.date.full],
        # *[
            # ('Project ' + k.title(), v) for k, v in vars(REPORT.project).items() if k != 'alias'
        # ]
    # ]:
        # print(key, ':', val)
    # print(DUMMY)
    
    # exit(0)
    # DUMMY = get_data()
    
    
    
    
    # REPORT = to_obj(get_data())
    REPORT = get_data()
    
    # for k in REPORT["headers"].keys():
        # h_obj = REPORT["headers"][k]
        # print(h_obj, type(h_obj), len(h_obj))
    
    import json
    with open("output.json", "w", encoding="utf-8") as f:
        json.dump(REPORT, f, ensure_ascii=False, indent=4)
    print('Json data saved.')
    exit(0)
    
    SUBJECT, RADIUS, LANG = f"""{REPORT["project"]["code"]} {REPORT["alias"]} as of {REPORT["date"]["value"]}""", "border-radius:6px", 'lang="en" dir="ltr"'
    HTML_PROPS = obj(
        wrapper=f"""{LANG} style="background-color:{PROJECT_DATA["colors"]["bg_dark"]};margin:0px;padding:0px;border:none;outline:none;width:100%;height:auto;-webkit-text-size-adjust:100%;-ms-text-size-adjust:100%;font-family:Ubuntu, Helvetica, Arial, sans-serif;">""",
        valign='valign="middle" style="vertical-align:middle', border=f"""border:0.169px solid {PROJECT_DATA["colors"]["fg"]}""", div_content=f"""<div style="dir:ltr;text-align:left;word-wrap:break-word;color:{PROJECT_DATA["colors"]["fg"]};font-size:16px;line-height:1.5;font-weight:300""",
        container=f"{RADIUS};width:512px;max-width:512px;background-color:#fff;margin:8px auto;padding:0px 24px 24px"
    )
    
    # labels, headers = [], []
    # for k in HEADER_LABELS:
        # print(k)
        # labels.append(k)
        # headers.append(v)
    
    header_div = f'<div {HTML_PROPS.valign} style="display:inline-block;line-height:0;font-size:15px;"'
    print(
        header_div + f'color:{PROJECT_DATA["colors"]["fg_var"]};">' + '\n\t'.join(labels) + '\n</div>',
        header_div + 'font-weight:600;">' + '\n'*3, '\n\t'.join(headers) + '\n</div>', '\n'*3
    )
    
    exit(0)
    
    def make_img(src='zee'):
        if src not in {'zee', 'hcc'}:
            return None
        defaults = f'{HTML_PROPS.valign};color:{PROJECT_DATA["colors"]["fg_var"]};height:auto;font-size:8px;line-height:1;text-decoration:none;outline:none;border:none'
        href, alt, size, filename = REPORT["zee"]["website"], REPORT["zee"]["name"], 64, 'cimacio'
        if src == 'hcc':
            href, alt, size, filename = REPORT.company.website, REPORT["company"]["name"], 32, 'hcc-logo'
        else:
            defaults += ";border-radius:100%"
        result = f'<a href="https://{href}/" width="{size}" {defaults};display:inline-block;width:{size}px;max-width:{size}px;" target="_blank">'
        result += f'<img src="https://raw.githubusercontent.com/MasterZeeno/SSR/refs/heads/main/{REPORT["zee"]["assets"]}/{filename}.png" alt="{alt}" {defaults};width:100%;max-width:100%;"></a>'
        return result
    
    def hr(s=12, c=3, a='center', v=0.69, f=PROJECT_DATA["colors"]["fg_var"]):
        a = a if a in ('center', 'left', 'right') else 'center'
        c, s, v = fmt_num(c, 1), fmt_num(s, 12), fmt_num(v, 0.69)
        return f"""<p role="separator" {HTML_PROPS.valign};user-select:none;cursor:none;color:{f};font-size:{s}px;opacity:{v};text-align:{a};">{'&mdash;'*c}</p>"""
        
    def br(c=1):
        c = fmt_num(c, 1)
        return '<br aria-hidden="true">' * c
    
    def bold(text=None, revert=False):
        prop = f"font-weight:{300 if revert else 600}"
        return f"""<span style="{prop};">{text}</span>""" if text else prop
    
    # def build_header():
        # label_block, main_block = [
            # (f"{s}color:{PROJECT_DATA["colors"]["fg_var"]}" if i == 0
            # else f"{s}font-weight:600") + ';">'
            # for i, s in enumerate(
                # [f'<div {HTML_PROPS.valign} style="display:inline-block;line-height:0;font-size:15px;"'] * 2
            # )
        # ]
        # for k, v in REPORT.headers:
            # label_block, main_block += [
                # *['<p style="line-height:0.5;">'] * 2,
                # *[k, v],
                # *['</p>']
            # ]
            
        
                
        
        
        
    summary_table = f"""<div style="padding:12px 0px;"><table width="100%" role="table" border="0" cellspacing="0" cellpadding="0" {HTML_PROPS.valign};border-collapse:collapse;width:100%;font-size:13px;color:{PROJECT_DATA["colors"]["fg"]};"><tbody>"""
    tag, consts = "td", f'{HTML_PROPS.valign};padding:3px 6px'
    for i, row in enumerate(REPORT.summary):
        row_html = "<tr>"
        for j, cell in enumerate(row):
            str_condition = True if i > 0 and j == 0 else False
            bgf_condition = True if i == 0 or (i in (8,9) and j == 3) else False
            align = "text-align:" + ("center" if (i == 0 and j == 0) else ("left" if j == 0 else "center"))
            bg = "background-color:" + (PROJECT_DATA["colors"]["bg"] if bgf_condition else "transparent")
            font_weight = bold() if bgf_condition else bold(revert=True)
            font_style = "font-style:" + ("italic" if str_condition else "normal")
            if str_condition:
                cell = "&nbsp;" * 3 + cell
            cell_html = f"""<{tag} align="{align}" {consts};{align};{HTML_PROPS.border};{bg};{font_weight};{font_style};">{cell}</{tag}>"""
            row_html += cell_html
        summary_table += row_html + "</tr>"
    
    return SUBJECT, f"""<!DOCTYPE html>
<html {LANG}>
<head>
    <meta charset="utf-8">
    <meta http-equiv="X-UA-Compatible" content="IE=edge">
    <title>{SUBJECT}</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
</head>
<body {HTML_PROPS.wrapper}
    <div {HTML_PROPS.wrapper}
        <div style="{HTML_PROPS.container};">
            <div style="padding:12px 24px 24px;">
                {HTML_PROPS.div_content};line-height:0;">
                    <p style="font-size:22px;{bold()};color:{PROJECT_DATA["colors"]["fg_lite"]}">{REPORT.title} ({REPORT["alias"]})</p>
                    {build_header()}
                </div>{br(2)}{hr(6, 69)}
                {HTML_PROPS.div_content};padding-top:16px;">
                    <p style="{bold()};font-size:18px;">{REPORT["email"]["msgs"][0]}</p><p>{REPORT["email"]["msgs"][1]}</p><p>{REPORT["email"]["msgs"][2]}</p>
                    {summary_table}</tbody></table></div><p>{REPORT["email"]["msgs"][3]}{bold(REPORT["email"]["msgs"][4])}</p>
                </div>
                {HTML_PROPS.div_content};">
                    {hr(a='left', f=PROJECT_DATA["colors"]["fg"])}<p>Best Regards,</p>
                    <div style="line-height:0;">
                        {make_img("zee")}
                        <div {HTML_PROPS.valign};display:inline-block;line-height:0.23;padding-left:8px;">
                            <p style="color:{PROJECT_DATA["colors"]["fg_lite"]};{bold()};font-size:16px;">{REPORT["zee"]["name"]}</p>
                            <p style="font-size:14px;">{REPORT["zee"]["position"]}</p>
                            <p style="font-size:12px;color:{PROJECT_DATA["colors"]["fg_var"]};">{REPORT["zee"]["licenses"]}</p>
                        </div>
                    </div>
                </div>
            </div>
            {br(2)}
            {HTML_PROPS.div_content};{RADIUS};background-color:{PROJECT_DATA["colors"]["bg_dark"]};padding:12px 24px 24px;">
                <p style="color:{PROJECT_DATA["colors"]["fg_var"]};text-align:justify;font-size:12px;">
                    {bold("Disclaimer:")}
                    {br(2)}
                    This is an {bold("automated message.")} Please do not reply directly to this email.
                    {br(2)}
                    This email, including any attachments and previous correspondence in the thread, is {bold("confidential")} and intended solely for the designated recipient(s).
                    If you are not the intended recipient, you are hereby notified that any review, dissemination, distribution, printing, or copying of this message and its contents is strictly prohibited.
                    If you have received this email in error or have unauthorized access to it, please notify the sender immediately and {bold("permanently delete all copies")} from your system.
                    {br(2)}
                    The sender and the organization shall not be held liable for any unintended transmission of confidential or privileged information.
                    {br(3)}
                </p>
                <div style="color:{PROJECT_DATA["colors"]["fg_var"]};text-align:center;">
                    {hr(6, 69)}{br()}{make_img("hcc")}
                    <div {HTML_PROPS.valign};display:inline-block;padding:4px 0px;line-height:0;">
                        <p style="font-size:16px;{bold()};color:{PROJECT_DATA["colors"]["fg"]};">{REPORT["company"]["name"]}</p>
                        <p style="font-size:9px;padding-bottom:6px;">{REPORT["company"]["site"]}</p>
                    </div>
                    <p style="font-size:10px;">{REPORT["company"]["licenses"]}{br()}{REPORT["company"]["copyleft"]}</p>
                </div>
            </div>
        </div>
    </div>
</body>
</html>"""


print(get_html_content())
