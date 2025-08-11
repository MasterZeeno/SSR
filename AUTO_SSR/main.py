import os
import re
import smtplib
import sys
from pathlib import Path
from datetime import datetime, date
from mimetypes import guess_type
from email.utils import formataddr
from openpyxl import load_workbook, Workbook
from email.message import EmailMessage
from calendar import month_name, month_abbr
from urllib.parse import urlparse, urlunparse, quote, parse_qsl, urlencode
from typing import Optional, Union, Dict
from types import SimpleNamespace

# --- CONSTANTS ---
SCRIPT_DIR = Path(__file__).resolve().parent

OPERATION: str = sys.argv[1] if len(sys.argv) > 1 else ""

MONTH_MAP: Dict[str, str] = {
    full: abbr for full, abbr in zip(month_name[1:], month_abbr[1:])
}

DATE_REGEX: str = r'\b(' + '|'.join(
    sorted(map(re.escape, MONTH_MAP), key=len, reverse=True)
) + r')\b'

DATE_RE = re.compile(
    (r"(\w+)\s+(\d{1,2})\s*-\s*(?:" +
    r"(\w+)\s+)?(\d{1,2}),?\s*(\d{4})"),
    flags=re.IGNORECASE
)

# --- FUNCTIONS WITH TYPING ---

def fmt_date(date_string: str) -> str:
    return re.sub(
        DATE_REGEX,
        lambda m: MONTH_MAP[m.group(0)],
        date_string
    )

def extract_dates(date_string):
    if not date_string or not date_string.strip():
        return None
    
    def parse(m, d, y):
        for fmt in (fmts := ("%B %d %Y", "%b %d %Y")):
            try:
                return SimpleNamespace(**{
                    "date": (dt := datetime.strptime(f"{m} {d} {y}", fmt).date()),
                    "string": SimpleNamespace(**{
                        k: quote(d) if "q" in k else d
                        for k, v in zip(
                            ["short", "full", "quoted"],
                            [f"{dt:%b}", [f"{dt:%B}"]*2]
                        )
                        if (d := f"{v} {dt.day}, {dt:%Y}")
                    })
                })
            except ValueError:
                continue
        return None
    
    match = DATE_RE.search(date_string).groups()
    
    if not match:
        return None
    
    return SimpleNamespace(**{
        key: [
            parse((m if m else match[0]), d, match[-1])
            for m, d in zip(match[::2], match[1::2])
        ]
        for key in ["start", "end"]
    })

def is_report_date(date_string: str, today=None) -> bool:
    if today is None:
        today = datetime.today().date()
        
    dates = extract_dates(date_string)
    return (
        False if dates is None else
        dates.end[0].date <= today
    )

def rslv_dir(dirname: Union[str, Path], parentdir: Optional[Union[str, Path]] = None) -> Path:
    base: Path = Path(parentdir) if parentdir else Path.cwd()
    directory: Path = (base / dirname).resolve()
    directory.mkdir(parents=True, exist_ok=True)
    return directory

def rel_to(filepath: Union[str, Path], basepath: Optional[Union[str, Path]] = None) -> str:
    filepath = Path(filepath).resolve()
    basepath = Path(basepath).resolve() if basepath else Path.cwd()
    try:
        return str(filepath.relative_to(basepath))
    except ValueError:
        return str(filepath)

def urlify(url: Union[str, bytes]) -> str:
    parsed = urlparse(url)
    encoded_path = quote(parsed.path, safe="")  # encode everything including slashes
    encoded_query = urlencode(parse_qsl(parsed.query), doseq=True)
    return urlunparse((
        parsed.scheme,
        parsed.netloc,
        encoded_path,
        parsed.params,
        encoded_query,
        parsed.fragment
    ))

def minify(text: str) -> str:
    text = re.sub(r'\n+', '', text.strip())
    text = re.sub(r'\s+', ' ', text)
    return re.sub(r'>\s+<', '><', text)

def send_email(subject: str, html_body: str, excel_file: Union[str, Path]) -> None:
    ZEE = (
        "Jay Ar Adlaon Cimacio, RN",
        "zeenoliev@gmail.com"
    )

    CFG: Dict[str, str] = {
        "Subject": subject,
        "From": formataddr(ZEE)
    }

    if "force" in OPERATION.lower():
        to_list = [
            f"jojofundales@{e}.com" +
            (".ph" if e == "hcc" else "")
            for e in ["hcc", "yahoo"]
        ]

        cc_yahoo = [f"{user}@yahoo.com" for user in [
            "arch_rbporral", "glachel.arao", "rbzden"
        ]]

        cc_gmail = [f"{user}@gmail.com" for user in [
            "maravilladarwin87.dm", "aljonporcalla",
            "eduardo111680"
        ]]

        CFG.update({
            "To": ", ".join(to_list),
            "Cc": ", ".join(cc_yahoo + cc_gmail)
        })
    else:
        CFG.update({
            "To": "cimaciojay0@gmail.com",
            "Cc": "yawapisting7@gmail.com"
        })

    msg = EmailMessage()
    for k, v in CFG.items():
        msg[k] = v

    msg.set_content("Greetings! ✨\n\nPlease see the attached file regarding the subject mentioned above.")
    msg.add_alternative(html_body, subtype="html")

    mime_type, _ = guess_type(str(excel_file))
    maintype, subtype = mime_type.split("/") if mime_type else ("application", "octet-stream")

    with open(excel_file, 'rb') as f:
        file_data = f.read()
        file_name = os.path.basename(excel_file)
        msg.add_attachment(
            file_data,
            maintype=maintype,
            subtype=subtype,
            filename=file_name
        )

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.starttls()
            smtp.login(ZEE[1], "frmoyroohmevbgvb")
            smtp.send_message(msg)
            print("Email sent successfully.")
    except Exception as e:
        print(f"Error sending email: {e}")
        
        
        
        
IMGS_DIR, WB_DIR = [
    rslv_dir(f"assets/{v}", SCRIPT_DIR)
    for v in ["imgs", "wb"]
]

WB_PATH = None
TODAY = datetime.today().date()

for excel_file in sorted(
    WB_DIR.glob('*.xlsx'),
    key=lambda f: f.stat().st_mtime, reverse=True
):
    if is_report_date(excel_file.stem, TODAY):
        WB_PATH = excel_file
        break

if WB_PATH:
    wb = load_workbook(WB_PATH, read_only=True, data_only=True)
    ws = [s for s in wb.worksheets if s.sheet_state == "visible"][-1]
    
    raw_data = {
        **{
            f"{p.name}_dir": urlify(r)
            for p in [IMGS_DIR, WB_DIR]
            if (r := rel_to(p, SCRIPT_DIR.parent))
        },
        "excel_file": urlify(WB_PATH.name),
        "year_now": str(datetime.now().year),
        "reference_no": str(ws.cell(61, 2).value).split()[-1],
        "report_period": fmt_date(str(ws.cell(63, 4).value))
    }
    
    for r in range(59, 68):
        for c in range(18, 21):
            val = ws.cell(r, c).value
            if val is not None:
                key = f"{r-59}:{c-18}"
                raw_data[key] = f"{val:,.0f}" if isinstance(val, (int, float)) else val
    
    data = {
        f"{{{{ {k} }}}}": v
        for k, v in raw_data.items()
    }
    
    subject = " ".join([
        "Update:", "PE-01-NSBP2-23",
        "Safety Statistics Report",
        "as of", raw_data["report_period"]
    ])
    
    template_path = Path("template.html").resolve()
    
    if template_path.exists() and data:
        html_content = template_path.read_text(encoding="utf-8")
        
        for k, v in data.items():
            html_content = html_content.replace(k, v)
        
        html_content = minify(html_content)
        send_email(subject, html_content, WB_PATH)
        with open(Path("index.html").resolve(), "w", encoding="utf-8") as f:
            f.write(html_content)






