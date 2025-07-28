from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
import os
import shutil
import sys
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.reader.drawings")

wb_basename = "NSB-P2 SSR"
wb_filename = f"{wb_basename}"

if not os.path.exists(f"{wb_filename}.xlsx"):
   sys.exit(1)

# Load source workbook and select the sheet
wb = load_workbook(f"{wb_filename}.xlsx", data_only=True)

shts = [sht for sht in wb.worksheets if sht.sheet_state == 'visible']
if not shts:
    sys.exit(1)

ws = shts[-1]
if not ws:
     sys.exit(1)

# for cell in ws.iter_rows(min_row=54, max_col=column_index_from_string("O"), values_only=True):
#     print(cell)

for row in range(54, ws.max_row + 1):
    for col in range(1, int(column_index_from_string("O")) + 1):
        value = str(ws.cell(row, col).value).strip()
        if value not in ('', 'None'):

sys.exit(0)

report_date = str(ws['Q56'].value).strip()
if report_date in ('', 'None'):
    sys.exit(1)

new_wb_filename = f"{wb_basename} as of {report_date}"
if os.path.exists(f"{new_wb_filename}.xlsx"):
    os.remove(f"{new_wb_filename}.xlsx")

# for row in ws.row_dimensions:
#     if ws.row_dimensions[row].hidden:
#         ws.row_dimensions[row].hidden = False

# ws.delete_rows(1, 53)

# ws.print_area = f"A1:N{str(ws.max_row)}"

# Delete columns beyond "O"
# Work backwards to avoid shifting issues
# for col in range(ws.max_column, column_index_from_string('O'), -1):
#     ws.delete_cols(col)

# for row in ws.row_dimensions:
#     if ws.row_dimensions[row].hidden:
#         ws.row_dimensions[row].hidden = False

# ws.delete_rows(1, 53)

# ws.print_area = f"A1:N{str(ws.max_row)}"

for sht in wb.worksheets:
    if sht != ws:
        wb.remove(sht)

ws.title = report_date
wb.save(f"{new_wb_filename}.xlsx")

wb = load_workbook(f"{new_wb_filename}.xlsx", data_only=True)
ws = wb.active

# Copy values up to column 'O'
# max_col = column_index_from_string("O")

for row in ws.iter_rows(values_only=True):
    for cell in row[:column_index_from_string("O")]:  # limit to column O
        new_cell = ws.cell(row=cell.row, column=cell.column)
        new_cell.value = cell.value

# for row in ws.iter_rows(values_only=True):
#     new_ws.append(row)

# Save as a copy
wb.save()

# max_col = column_index_from_string("O")

# for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
#     for cell in row[:max_col]:  # limit to column O
#         new_cell = ws.cell(row=cell.row, column=cell.column)
#         new_cell.value = cell.value

sys.exit(0)
